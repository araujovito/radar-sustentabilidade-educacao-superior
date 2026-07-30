"""Perfil estrutural do pacote oficial do Censo Superior."""

import csv
import hashlib
import json
import re
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import openpyxl

from radar_sustentabilidade.ingestion.download import sha256_file

MD5_PATTERN = re.compile(r"^(?P<hash>[0-9a-fA-F]{32})\s+\*(?P<name>.+)$")


def _decode_sample(sample: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            return sample.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("unknown", sample, 0, 1, "encoding not detected")


def _parse_md5_control(content: str) -> list[dict]:
    entries = []
    for line in content.splitlines():
        match = MD5_PATTERN.match(line.strip())
        if match:
            entries.append(
                {
                    "md5": match.group("hash").lower(),
                    "file_name": match.group("name"),
                }
            )
    return entries


def _csv_member_profile(archive: ZipFile, member_name: str) -> dict:
    digest = hashlib.md5()
    newline_count = 0
    total_bytes = 0
    last_byte = b""

    with archive.open(member_name) as stream:
        sample = stream.read(128 * 1024)
        digest.update(sample)
        newline_count += sample.count(b"\n")
        total_bytes += len(sample)
        if sample:
            last_byte = sample[-1:]

        while chunk := stream.read(8 * 1024 * 1024):
            digest.update(chunk)
            newline_count += chunk.count(b"\n")
            total_bytes += len(chunk)
            last_byte = chunk[-1:]

    line_count = newline_count + int(bool(total_bytes) and last_byte != b"\n")
    text, encoding = _decode_sample(sample)
    first_line = text.splitlines()[0]
    delimiter = csv.Sniffer().sniff(first_line, delimiters=";,|\t").delimiter
    columns = next(csv.reader([first_line], delimiter=delimiter))

    return {
        "member_path": member_name,
        "file_name": Path(member_name).name,
        "file_size_bytes": total_bytes,
        "md5": digest.hexdigest(),
        "line_count": line_count,
        "data_row_count": max(line_count - 1, 0),
        "encoding": encoding,
        "delimiter": delimiter,
        "column_count": len(columns),
        "columns": columns,
    }


def _dictionary_profile(content: bytes) -> dict:
    workbook = openpyxl.load_workbook(
        BytesIO(content),
        read_only=True,
        data_only=True,
    )
    sheets = []
    for worksheet in workbook.worksheets:
        variable_names = [
            value
            for (value,) in worksheet.iter_rows(
                min_row=9,
                min_col=2,
                max_col=2,
                values_only=True,
            )
            if isinstance(value, str) and value.strip()
        ]
        sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "variable_count": len(variable_names),
                "variables": variable_names,
            }
        )
    return {"sheets": sheets}


def profile_package(archive_path: Path) -> dict:
    """Perfila membros, tabelas, hashes oficiais e dicionário do ZIP."""
    with ZipFile(archive_path) as archive:
        names = archive.namelist()
        real_names = [
            name for name in names if not Path(name).name.startswith("~$")
        ]
        csv_names = [name for name in real_names if name.lower().endswith(".csv")]
        xlsx_names = [name for name in real_names if name.lower().endswith(".xlsx")]
        txt_names = [name for name in real_names if name.lower().endswith(".txt")]

        if len(xlsx_names) != 1:
            raise ValueError(
                f"Esperado um dicionário XLSX; encontrados {len(xlsx_names)}"
            )
        if len(txt_names) != 1:
            raise ValueError(
                f"Esperado um controle MD5; encontrados {len(txt_names)}"
            )

        control_text, control_encoding = _decode_sample(
            archive.read(txt_names[0])
        )
        md5_entries = _parse_md5_control(control_text)
        tables = [
            _csv_member_profile(archive, member_name)
            for member_name in csv_names
        ]
        dictionary = _dictionary_profile(archive.read(xlsx_names[0]))

    expected_by_hash = {entry["md5"]: entry for entry in md5_entries}
    sheets_by_name = {sheet["name"].lower(): sheet for sheet in dictionary["sheets"]}

    for table in tables:
        control_entry = expected_by_hash.get(table["md5"])
        table["md5_matches_control"] = control_entry is not None
        table["control_file_name"] = (
            control_entry["file_name"] if control_entry else None
        )
        table["control_name_matches_member"] = bool(
            control_entry
            and control_entry["file_name"].lower() == table["file_name"].lower()
        )

        sheet_key = (
            "cadastro_cursos"
            if "CURSOS" in table["file_name"].upper()
            else "cadastro_ies"
        )
        dictionary_sheet = sheets_by_name.get(sheet_key)
        dictionary_variables = (
            dictionary_sheet["variables"] if dictionary_sheet else []
        )
        table["dictionary_sheet"] = sheet_key
        table["dictionary_variable_count"] = len(dictionary_variables)
        table["columns_missing_in_dictionary"] = sorted(
            set(table["columns"]) - set(dictionary_variables)
        )
        table["dictionary_variables_missing_in_csv"] = sorted(
            set(dictionary_variables) - set(table["columns"])
        )

    return {
        "schema_version": 1,
        "archive": {
            "file_name": archive_path.name,
            "file_size_bytes": archive_path.stat().st_size,
            "sha256": sha256_file(archive_path),
        },
        "member_count": len(names),
        "dictionary_member": xlsx_names[0],
        "md5_control_member": txt_names[0],
        "md5_control_encoding": control_encoding,
        "tables": tables,
        "dictionary": dictionary,
    }


def write_package_profile(archive_path: Path, output_path: Path) -> dict:
    """Gera e persiste o perfil estrutural do pacote."""
    profile = profile_package(archive_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(profile, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return profile
