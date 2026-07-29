import hashlib
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_STORED, ZipFile

import openpyxl

from radar_sustentabilidade.ingestion.profile import profile_package


def make_dictionary() -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "cadastro_cursos"
    worksheet.cell(row=9, column=2, value="NU_ANO_CENSO")
    worksheet.cell(row=10, column=2, value="CO_CURSO")
    worksheet.cell(row=11, column=2, value="QT_MAT")

    ies_sheet = workbook.create_sheet("cadastro_ies")
    ies_sheet.cell(row=9, column=2, value="NU_ANO_CENSO")
    ies_sheet.cell(row=10, column=2, value="CO_IES")

    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def test_profiles_official_package_structure(tmp_path: Path) -> None:
    archive_path = tmp_path / "microdados_2024.zip"
    courses = (
        "NU_ANO_CENSO;CO_CURSO;QT_MAT\n"
        "2024;10;100\n"
        "2024;20;200\n"
    ).encode("latin-1")
    institutions = "NU_ANO_CENSO;CO_IES\n2024;1\n".encode("latin-1")
    control = (
        f"{hashlib.md5(institutions).hexdigest()} "
        "*MICRODADOS_CADASTRO_IES_2024.csv\r\n"
        f"{hashlib.md5(courses).hexdigest()} "
        "*MICRODADOS_CADASTRO_CURSOS_2024.csv\r\n"
    )

    with ZipFile(archive_path, "w", compression=ZIP_STORED) as archive:
        archive.writestr(
            "Anexos/dicionario.xlsx",
            make_dictionary(),
        )
        archive.writestr(
            "dados/MICRODADOS_CADASTRO_CURSOS_2024.CSV",
            courses,
        )
        archive.writestr(
            "dados/MICRODADOS_ED_SUP_IES_2024.CSV",
            institutions,
        )
        archive.writestr("dados/md5.txt", control)

    profile = profile_package(archive_path)
    tables = {table["dictionary_sheet"]: table for table in profile["tables"]}

    assert profile["member_count"] == 4
    assert tables["cadastro_cursos"]["data_row_count"] == 2
    assert tables["cadastro_cursos"]["column_count"] == 3
    assert tables["cadastro_cursos"]["md5_matches_control"] is True
    assert tables["cadastro_cursos"]["control_name_matches_member"] is True
    assert tables["cadastro_cursos"]["columns_missing_in_dictionary"] == []
    assert tables["cadastro_ies"]["control_name_matches_member"] is False
